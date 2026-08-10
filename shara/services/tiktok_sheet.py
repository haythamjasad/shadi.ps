from __future__ import annotations

import hashlib
import re
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional

import requests
from openpyxl import load_workbook


def _root_dir() -> Path:
    return Path(__file__).resolve().parent.parent


def _sheet_path(raw_path: str | None) -> Path:
    if raw_path:
        candidate = Path(raw_path).expanduser().resolve()
        if candidate.exists():
            return candidate
    root_parent = _root_dir().parent
    default = root_parent / "tiktok_videos_shadishirri.xlsx"
    if default.exists():
        return default.resolve()
    matches = sorted(root_parent.glob("*tiktok*.xlsx"))
    if matches:
        return matches[0].resolve()
    return default.resolve()


def _normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _reel_id_from_url(url: str) -> str:
    match = re.search(r"/video/(\d+)", str(url or ""))
    if match:
        return match.group(1)
    return hashlib.sha1(str(url or "").encode("utf-8", errors="ignore")).hexdigest()


def _download_thumbnail(thumbnail_url: str, reel_id: str) -> str:
    source = _normalize_text(thumbnail_url)
    if not source:
        return ""

    if source.startswith("/images/") or source.startswith("images/"):
        return source if source.startswith("/") else f"/{source.lstrip('/')}"

    target_dir = _root_dir() / "images" / "reel_thumbnails"
    target_dir.mkdir(parents=True, exist_ok=True)

    ext = Path(source.split("?", 1)[0]).suffix or ".jpg"
    target_path = target_dir / f"{reel_id}{ext}"
    if target_path.exists() and target_path.is_file():
        return f"/images/reel_thumbnails/{target_path.name}"

    response = requests.get(
        source,
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=45,
    )
    response.raise_for_status()
    content_type = response.headers.get("Content-Type", "")
    if "png" in content_type:
        target_path = target_dir / f"{reel_id}.png"
    elif "webp" in content_type:
        target_path = target_dir / f"{reel_id}.webp"
    elif "gif" in content_type:
        target_path = target_dir / f"{reel_id}.gif"
    elif "jpeg" in content_type or "jpg" in content_type:
        target_path = target_dir / f"{reel_id}.jpg"

    target_path.write_bytes(response.content)
    return f"/images/reel_thumbnails/{target_path.name}"


@lru_cache(maxsize=1)
def _load_rows(sheet_path: str) -> tuple[Dict[str, Any], ...]:
    wb = load_workbook(_sheet_path(sheet_path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    total = len(rows)

    items: list[Dict[str, Any]] = []
    for index, row in enumerate(rows):
        reel_url = _normalize_text(row[0] if len(row) > 0 else "")
        thumbnail_url = _normalize_text(row[1] if len(row) > 1 else "")
        upload_date = _normalize_text(row[2] if len(row) > 2 else "")
        title = _normalize_text(row[4] if len(row) > 4 else "")
        if not reel_url or not title:
            continue

        reel_id = _reel_id_from_url(reel_url)
        try:
            local_thumb = _download_thumbnail(thumbnail_url, reel_id)
        except Exception:
            local_thumb = ""
        items.append(
            {
                "id": reel_id,
                "platform": "tiktok",
                "facebookReelUrl": reel_url,
                "title": title,
                "thumbnail": local_thumb or thumbnail_url,
                "uploadDate": upload_date,
                "createdAt": float(total - index),
                "popularityScore": 0,
                "tags": [],
                "topic": "عام",
                "duration": "",
            }
        )

    return tuple(items)


def list_tiktok_sheet_reels(*, sheet_path: str, limit: Optional[int] = 100) -> List[Dict[str, Any]]:
    reels = list(_load_rows(sheet_path))
    if limit is None:
        return reels
    return reels[: max(1, int(limit))]
